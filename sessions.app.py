import streamlit as st
import pandas as pd
import datetime
import io
from openpyxl.styles import PatternFill

# ---------- Initial Setup ----------
st.set_page_config(page_title="Therapy Session Tracker", layout="centered")

# ---------- Aging Bucket Function ----------
def aging_bucket(days):
    if days <= 30:
        return "0-30 days"
    elif days <= 60:
        return "31-60 days"
    elif days <= 90:
        return "61-90 days"
    else:
        return "90+ days"

# ---------- Initialize in-memory DataFrame ----------
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame(columns=[
        "Clinician", "Client Initials", "Date of Service", "CPT Code", "Session Fee",
        "Payment Received", "Date of Payment", "Outstanding",
        "Days Outstanding", "Aging Bucket"
    ])

df = st.session_state.df

# ---------- Clear Session Data ----------
if not df.empty:
    if st.button("🗑️ Clear All Session Data"):
        st.session_state.df = pd.DataFrame(columns=df.columns)
        st.success("✅ All in-browser session data cleared!")

# ---------- Warning for unsaved data ----------
if not df.empty:
    st.warning("⚠️ You have unsaved session data in this browser session. "
               "Make sure to download your Excel file before closing or refreshing!")

# ---------- Excel Export with Color ----------
def export_colored_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sessions")
        ws = writer.sheets["Sessions"]
        
        # Find the Aging Bucket column
        aging_col_idx = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Aging Bucket":
                aging_col_idx = idx
                break
        
        color_map = {
            "Paid": "00C6EFCE",
            "0-30 days": "00C6EFCE",
            "31-60 days": "00FFEB9C",
            "61-90 days": "00F4B084",
            "90+ days": "00FF0000"
        }
        
        if aging_col_idx:
            for row in ws.iter_rows(min_row=2, min_col=aging_col_idx, max_col=aging_col_idx):
                cell = row[0]
                fill_color = color_map.get(cell.value, None)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    output.seek(0)
    return output

# ---------- Input Form ----------
st.title("📊 Therapy Session Tracker (Persistent in-browser)")

with st.form("session_entry"):
    clinician = st.text_input("Clinician Name")
    client_initials = st.text_input("Client Initials")
    date_of_service = st.date_input("Date of Service", datetime.date.today())
    cpt_code = st.selectbox("CPT Code", ["90837", "90791"])
    session_fee = st.number_input("Session Fee ($)", min_value=0.0, step=10.0)
    payment_received = st.number_input("Payment Received ($)", min_value=0.0, step=10.0)

    unpaid = st.checkbox("Unpaid?")
    if unpaid:
        date_of_payment = None
    else:
        date_of_payment = st.date_input("Date of Payment", datetime.date.today())

    submitted = st.form_submit_button("Add Session")

if submitted:
    outstanding = session_fee - payment_received
    if unpaid or not date_of_payment:
        days_outstanding = (datetime.date.today() - date_of_service).days if outstanding > 0 else 0
    else:
        days_outstanding = (datetime.date.today() - date_of_payment).days if outstanding > 0 else 0

    bucket = aging_bucket(days_outstanding) if outstanding > 0 else "Paid"

    new_row = {
        "Clinician": clinician,
        "Client Initials": client_initials,
        "Date of Service": date_of_service,
        "CPT Code": cpt_code,
        "Session Fee": session_fee,
        "Payment Received": payment_received,
        "Date of Payment": date_of_payment,
        "Outstanding": outstanding,
        "Days Outstanding": days_outstanding,
        "Aging Bucket": bucket
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    st.session_state.df = df
    st.success("✅ Session added (in-browser memory)")

# ---------- Edit Existing Session ----------
if not df.empty:
    st.sidebar.header("✏️ Edit Existing Session")
    row_to_edit = st.sidebar.selectbox("Select row to edit (by index)", df.index.tolist())
    if row_to_edit is not None:
        with st.sidebar.form("edit_form"):
            clinician_edit = st.text_input("Clinician Name", value=df.at[row_to_edit, "Clinician"])
            client_edit = st.text_input("Client Initials", value=df.at[row_to_edit, "Client Initials"])
            date_edit = st.date_input("Date of Service", value=pd.to_datetime(df.at[row_to_edit, "Date of Service"]))

            cpt_options = ["90837", "90791"]
            try:
                default_index = cpt_options.index(df.at[row_to_edit, "CPT Code"])
            except ValueError:
                default_index = 0
            cpt_edit = st.selectbox("CPT Code", cpt_options, index=default_index)

            fee_edit = st.number_input("Session Fee ($)", min_value=0.0, value=float(df.at[row_to_edit, "Session Fee"]), step=1.0)
            payment_edit = st.number_input("Payment Received ($)", min_value=0.0, value=float(df.at[row_to_edit, "Payment Received"]), step=1.0)

            unpaid_edit = st.checkbox("Unpaid?", value=df.at[row_to_edit, "Outstanding"]>0 and pd.isna(df.at[row_to_edit,"Date of Payment"]))
            if unpaid_edit:
                date_payment_edit = None
            else:
                date_payment_edit = st.date_input(
                    "Date of Payment (optional)",
                    value=pd.to_datetime(df.at[row_to_edit, "Date of Payment"]) if not pd.isna(df.at[row_to_edit, "Date of Payment"]) else datetime.date.today()
                )

            save_edit = st.form_submit_button("Save Changes")

            if save_edit:
                df.at[row_to_edit, "Clinician"] = clinician_edit
                df.at[row_to_edit, "Client Initials"] = client_edit
                df.at[row_to_edit, "Date of Service"] = pd.to_datetime(date_edit)
                df.at[row_to_edit, "CPT Code"] = cpt_edit
                df.at[row_to_edit, "Session Fee"] = fee_edit
                df.at[row_to_edit, "Payment Received"] = payment_edit
                df.at[row_to_edit, "Date of Payment"] = pd.to_datetime(date_payment_edit) if date_payment_edit else pd.NaT
                df.at[row_to_edit, "Outstanding"] = df.at[row_to_edit, "Session Fee"] - df.at[row_to_edit, "Payment Received"]
                if unpaid_edit or not date_payment_edit:
                    df.at[row_to_edit, "Days Outstanding"] = (datetime.date.today() - df.at[row_to_edit, "Date of Service"].date()).days if df.at[row_to_edit, "Outstanding"]>0 else 0
                else:
                    df.at[row_to_edit, "Days Outstanding"] = (datetime.date.today() - df.at[row_to_edit, "Date of Payment"].date()).days if df.at[row_to_edit, "Outstanding"]>0 else 0
                df.at[row_to_edit, "Aging Bucket"] = aging_bucket(df.at[row_to_edit, "Days Outstanding"]) if df.at[row_to_edit, "Outstanding"]>0 else "Paid"
                st.session_state.df = df
                st.success("✅ Session updated (in-browser memory)")

# ---------- Ensure proper datetime ----------
for col in ["Date of Service", "Date of Payment"]:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")

# ---------- Display Data ----------
st.subheader("📋 All Sessions")
st.dataframe(df, use_container_width=True)

# ---------- Aging Summary ----------
if not df.empty:
    st.subheader("📊 Aging Summary")
    summary = df.groupby("Aging Bucket")["Outstanding"].sum().reset_index()
    color_map = {
        "0-30 days": "🟩",
        "31-60 days": "🟨",
        "61-90 days": "🟧",
        "90+ days": "🟥",
        "Paid": "✅"
    }
    summary["Status"] = summary["Aging Bucket"].map(color_map)
    st.table(summary[["Status", "Aging Bucket", "Outstanding"]])

# ---------- Download Excel ----------
if not df.empty:
    st.subheader("⬇️ Download Data")
    excel_bytes = export_colored_excel(df)
    st.download_button(
        label="📥 Download Excel with Colors",
        data=excel_bytes,
        file_name="sessions.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
